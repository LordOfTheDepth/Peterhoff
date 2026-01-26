import os
import json
import shutil
from openpyxl import load_workbook
import logging
from datetime import datetime
from ThumbnailsGenerator import ThumbnailGenerator
import DocxConverter
import CategorizerUtills

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_thumbnails_for_folder(folder_path, thumbnail_generator):
    """
    Создает миниатюры для всех изображений в указанной папке
    
    Args:
        folder_path (str): Путь к папке с изображениями
        thumbnail_generator: Экземпляр ThumbnailGenerator
    
    Returns:
        int: Количество созданных миниатюр
    """
    # Создаем папку для миниатюр
    thumbnails_folder = os.path.join(folder_path, "thumbnails")
    os.makedirs(thumbnails_folder, exist_ok=True)
    
    created_count = 0
    
    try:
        # Получаем список всех файлов в папке
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            
            # Проверяем, является ли файл изображением
            if os.path.isfile(file_path):
                ext = os.path.splitext(filename)[1].lower()
                image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp'}
                
                if ext in image_extensions:
                    # Создаем миниатюру
                    result = thumbnail_generator.createThumbnail(file_path, thumbnails_folder)
                    if result:
                        created_count += 1
                        logger.debug(f"Создана миниатюра для: {filename}")
        
        if created_count > 0:
            logger.info(f"Создано {created_count} миниатюр в {thumbnails_folder}")
        else:
            logger.info(f"В папке нет изображений для создания миниатюр: {folder_path}")
            
    except Exception as e:
        logger.error(f"Ошибка при создании миниатюр для папки {folder_path}: {e}")
    
    return created_count

def find_excel_file_in_folder(source_folder):
    """
    Находит первый .xlsx файл в указанной папке и всех подпапках
    
    Args:
        source_folder (str): Путь к исходной папке
    
    Returns:
        str or None: Путь к найденному Excel файлу или None
    """
    try:
        # Получаем все файлы рекурсивно
        all_files = CategorizerUtills.get_all_files_recursive(source_folder)
        
        # Ищем первый .xlsx файл
        for full_path, file_info in all_files.items():
            filename = file_info['filename']
            if filename.lower().endswith('.xlsx'):
                logger.info(f"Найден Excel файл: {full_path}")
                return full_path
        
        logger.warning(f"Не найден Excel файл (.xlsx) в папке: {source_folder}")
        return None
        
    except Exception as e:
        logger.error(f"Ошибка при поиске Excel файла: {e}")
        return None

def DoFolder(source_folder, dest_folder):
    # Преобразуем относительные пути в абсолютные
    source_folder_fixed = CategorizerUtills.resolve_relative_path(source_folder)
    dest_folder_fixed = CategorizerUtills.resolve_relative_path(dest_folder)

    thumbnail_generator = ThumbnailGenerator(
        max_size=(300, 300),
        quality=85,
        format='JPEG'
    )

    not_found_reports = []
    
    # Инициализируем структуру для map.json
    map_data = {
        "folders": {}
    }

    # Проверяем существование исходной папки
    if not os.path.exists(source_folder_fixed):
        logger.error(f"Исходная папка не найдена: {source_folder_fixed}")
        return
    
    # Автоматически находим Excel файл в исходной папке
    table_fixed = find_excel_file_in_folder(source_folder_fixed)
    if not table_fixed:
        logger.error(f"Excel файл не найден в папке: {source_folder_fixed}")
        return
    
    # Создаем конечную папку, если ее нет
    os.makedirs(dest_folder_fixed, exist_ok=True)
    
    # КОНВЕРТИРУЕМ DOCX ФАЙЛЫ ИЗ ИСХОДНОЙ ПАПКИ
    logger.info(f"Поиск docx файлов в исходной папке...")
    DocxConverter.convert_docx_in_folder(source_folder_fixed, dest_folder_fixed)

    # Загружаем все файлы из source_folder и всех подпапок
    logger.info(f"Поиск файлов в {source_folder_fixed} и подпапках...")
    all_files = CategorizerUtills.get_all_files_recursive(source_folder_fixed)
    
    # Создаем карту нормализованных имен
    normalized_dict = {}
    file_count = 0
    
    for full_path, file_info in all_files.items():
        filename = file_info['filename']
        # Пропускаем Excel файлы из карты нормализованных имен
        if filename.lower().endswith('.xlsx'):
            continue
            
        normalized_filename = CategorizerUtills.normalize_filename(filename)
        
        if normalized_filename not in normalized_dict:
            normalized_dict[normalized_filename] = []
        
        normalized_dict[normalized_filename].append({
            'title': "",
            "copyright": "",
            "description": "",
            'filename': filename,
            'full_path': full_path,
            'rel_path': file_info['rel_path']
        })
        file_count += 1
    
    # Логируем количество уникальных нормализованных имен
    logger.info(f"Найдено {file_count} файлов, {len(normalized_dict)} уникальных нормализованных имен")
    
    # Загружаем Excel файл
    try:
        wb = load_workbook(table_fixed, data_only=True)
        logger.info(f"Загружен Excel файл с листами: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Ошибка при загрузке Excel файла: {e}")
        return
    
    # Обрабатываем каждый лист
    for sheet_name in wb.sheetnames:
        
        logger.info(f"Обработка листа: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Получаем название папки из первой строки, первого столбца (столбец A)
        folder_name_cell = ws.cell(row=2, column=1).value
        if folder_name_cell:
            folder_name = str(folder_name_cell).strip()
        else:
            logger.error(f"Не найдена ячейка с заголовком на листе: {sheet_name}")
            folder_name = "   "
        
        # Получаем название подпапки из второй строки, второго столбца (столбец B)
        subfolder_name_cell = ws.cell(row=2, column=2).value
        if subfolder_name_cell:
            subfolder_name = str(subfolder_name_cell).strip()
        else:
            subfolder_name = ""
        
        logger.info(f"Папка: {folder_name}")
        if subfolder_name:
            logger.info(f"Подпапка: {subfolder_name}")
        
        # Инициализируем структуру для этой папки в map_data
        if folder_name not in map_data["folders"]:
            map_data["folders"][folder_name] = {
                "files": [],
                "subfolders": {}
            }
        
        # Если есть подпапка, инициализируем ее структуру
        if subfolder_name:
            if subfolder_name not in map_data["folders"][folder_name]["subfolders"]:
                map_data["folders"][folder_name]["subfolders"][subfolder_name] = {
                    "files": []
                }
        
        # Собираем статистику
        row_count = 0
        copied_count = 0
        not_found_count = 0
        multiple_found_count = 0
        
        # Список не найденных файлов для этого листа
        sheet_not_found = []
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Пропускаем пустые строки
            if not row or row[0] is None:
                continue
            
            image_name = str(row[0]).strip()
            description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            
            row_count += 1
            
            # Нормализуем имя для поиска
            normalized_table_entry = CategorizerUtills.normalize_filename(image_name)
            
            # Ищем совпадение файла
            if normalized_table_entry in normalized_dict:
                matched_files = normalized_dict[normalized_table_entry]
                
                if len(matched_files) > 1:
                    logger.warning(f"Найдено {len(matched_files)} файлов для '{image_name}': {[f['filename'] for f in matched_files]}")
                    multiple_found_count += 1
                
                # Берем первый найденный файл
                file_info = matched_files[0]
                matched_file_path = file_info['full_path']
                matched_filename = file_info['filename']
                
                # Подготовка данных для map.json
                file_data = {
                    "description": description,
                    "filename": matched_filename
                }
                
                # Добавляем файл в соответствующую структуру
                if subfolder_name:
                    # Файл в подпапке
                    map_data["folders"][folder_name]["subfolders"][subfolder_name]["files"].append(file_data)
                else:
                    # Файл в основной папке
                    map_data["folders"][folder_name]["files"].append(file_data)
                
                # Копируем файл в корень целевой папки
                dest_path = os.path.join(dest_folder_fixed, matched_filename)
                
                try:
                    # Проверяем, не существует ли уже файл
                    if os.path.exists(dest_path):
                        # Добавляем суффикс если файл уже существует
                        base, ext = os.path.splitext(matched_filename)
                        counter = 1
                        while os.path.exists(dest_path):
                            new_filename = f"{base}_{counter}{ext}"
                            dest_path = os.path.join(dest_folder_fixed, new_filename)
                            counter += 1
                        
                        final_filename = os.path.basename(dest_path)
                        logger.info(f"Файл '{matched_filename}' уже существует, сохранен как '{final_filename}'")
                        
                        # Обновляем имя файла в map_data
                        file_data["filename"] = final_filename
                    else:
                        final_filename = matched_filename
                    
                    shutil.copy2(matched_file_path, dest_path)
                    copied_count += 1
                    
                except Exception as e:
                    logger.error(f"Ошибка при копировании из {matched_file_path} в {dest_path}: {e}")
            else:
                logger.warning(f"Не найден файл для: '{image_name}' на листе '{sheet_name}'")
                
                # Сохраняем информацию о не найденном файле
                sheet_not_found.append({
                    'title': folder_name,
                    'subtitle': subfolder_name,
                    'name': image_name,
                    'sheet': sheet_name
                })
                
                # Для отладки: выводим похожие имена
                similar = [k for k in normalized_dict.keys() if normalized_table_entry in k or k in normalized_table_entry]
                if similar:
                    logger.debug(f"Похожие нормализованные имена: {similar[:5]}")
                
                not_found_count += 1
        
        # Добавляем не найденные файлы из этого листа в общий отчет
        if sheet_not_found:
            not_found_reports.extend(sheet_not_found)
        
        logger.info(f"Лист '{sheet_name}': обработано {row_count} строк, найдено {copied_count}, не найдено {not_found_count}, множественные совпадения: {multiple_found_count}")
    
    # СОЗДАЕМ МИНИАТЮРЫ ДЛЯ ЦЕЛЕВОЙ ПАПКИ
    if thumbnail_generator and os.path.exists(dest_folder_fixed):
        thumbnails_created = create_thumbnails_for_folder(dest_folder_fixed, thumbnail_generator)
        logger.info(f"Создано {thumbnails_created} миниатюр в папке: {dest_folder_fixed}")
    
    # СОЗДАЕМ MAP.JSON В КОНЕЧНОЙ ПАПКЕ
    json_path = os.path.join(dest_folder_fixed, "map.json")
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(map_data, f, ensure_ascii=False, indent=4)
        logger.info(f"Создан map.json в папке: {dest_folder_fixed}")
    except Exception as e:
        logger.error(f"Ошибка при создании map.json: {e}")

    # Сохраняем общий отчет о не найденных файлах в корневой целевой папке
    if not_found_reports:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"not_found_report_{timestamp}.txt"
        report_path = os.path.join(dest_folder_fixed, report_filename)
        
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"ОТЧЕТ О НЕ НАЙДЕННЫХ ФАЙЛАХ\n")
                f.write(f"===========================\n\n")
                f.write(f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
                f.write(f"Исходная папка: {source_folder_fixed}\n")
                f.write(f"Excel файл: {table_fixed}\n")
                f.write(f"Целевая папка: {dest_folder_fixed}\n")
                f.write(f"Всего не найденных файлов: {len(not_found_reports)}\n\n")
                f.write(f"СПИСОК НЕ НАЙДЕННЫХ ФАЙЛОВ:\n")
                f.write(f"============================\n\n")
                
                for item in not_found_reports:
                    f.write(f"Заголовок: {item['title']}\n")
                    f.write(f"Подзаголовок: {item['subtitle']}\n")
                    f.write(f"Название файла: {item['name']}\n")
                    f.write(f"Лист Excel: {item['sheet']}\n")
                    f.write(f"{'-'*50}\n")
            
            logger.info(f"Создан общий отчет о не найденных файлах: {report_path}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении общего отчета: {e}")
    else:
        logger.info("Все файлы успешно найдены. Отчет не требуется.")
    
    logger.info("Обработка завершена!")
    
    # Дополнительная диагностика
    logger.info("=== ДИАГНОСТИКА ===")
    logger.info(f"Всего файлов в исходной папке: {file_count}")
    logger.info(f"Уникальных нормализованных имен: {len(normalized_dict)}")

def main():
   
    sortedPath = CategorizerUtills.resolve_relative_path("SortedMap")
    if os.path.exists(sortedPath):
        try:
            shutil.rmtree(sortedPath)
            print(f"Directory '{sortedPath}' and all its contents deleted.")
        except OSError as error:
            print(f"Error: {error}")
    else:
        print(f"Directory '{sortedPath}' not found.")

    locations = ["Петергоф"]#,"Пушкин","Павловск"]
    stages = ["до войны","разрушения","восстановление"]
    
    for location in locations:
        for stage in stages:
            folderPath = f"{location}\\{stage}"
            DoFolder(
                f"Unsorted\\{folderPath}",
                f"SortedMap\\{folderPath}"
            )
        
if __name__ == "__main__":
    main()