import os
import json
import shutil
from openpyxl import load_workbook
import logging
from datetime import datetime

# Импортируем генератор миниатюр из того же каталога
try:
    from ThumbnailsGenerator import ThumbnailGenerator
    THUMBNAIL_GENERATOR_AVAILABLE = True
    logger = logging.getLogger(__name__)
    logger.info("ThumbnailGenerator успешно импортирован")
except ImportError as e:
    THUMBNAIL_GENERATOR_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning(f"Не удалось импортировать ThumbnailGenerator: {e}")
    logger.warning("Миниатюры создаваться не будут")

# ========== НАСТРОЙКИ ==========
THUMBNAIL_SIZE = (300, 300)  # Максимальный размер миниатюры
THUMBNAIL_QUALITY = 85       # Качество миниатюр (1-100)
# ===============================

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_base_dir():
    """Возвращает базовую директорию скрипта"""
    return os.path.dirname(os.path.abspath(__file__))

def resolve_relative_path(relative_path):
    """Преобразует относительный путь в абсолютный относительно директории скрипта"""
    base_dir = get_base_dir()
    # Если путь уже абсолютный, возвращаем его как есть
    if os.path.isabs(relative_path):
        return relative_path
    return os.path.join(base_dir, relative_path)

def normalize_filename(filename):
    """Удаляет расширения изображений и оставляет только буквы и цифры"""
    # Определяем расширения файлов изображений
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg'}
    
    # Находим расширение файла (последняя точка в имени)
    name, ext = os.path.splitext(filename)
    ext_lower = ext.lower()
    
    # Если это расширение изображения - удаляем его, иначе оставляем как есть
    if ext_lower in image_extensions:
        name_without_ext = name
    else:
        name_without_ext = filename  # Если не изображение, оставляем все имя
    
    # Создаем пустую строку для результата
    normalized_chars = []
    
    # Проходим по каждому символу
    for char in name_without_ext:
        if char.isalpha() or char.isdigit():
            normalized_chars.append(char)
    
    # Объединяем и приводим к нижнему регистру
    normalized = ''.join(normalized_chars)
    return normalized.lower()

def sanitize_folder_name(folder_name):
    """Очищает имя папки от недопустимых символов"""
    # Удаляем недопустимые символы для имен папок в Windows
    # <>:"/\|?* а также управляющие символы
    invalid_chars = '<>:"/\\|?*\n'
    for char in invalid_chars:
        folder_name = folder_name.replace(char, '')
    
    # Удаляем начальные и конечные пробелы и точки
    folder_name = folder_name.strip(' .')
    
    # Если после очистки имя пустое, используем имя листа
    if not folder_name:
        return "Без_названия"
    
    return folder_name

def ensure_descriptions_json_in_all_parents(folder_path, folder_name, subfolder_name=""):
    """Обеспечивает наличие descriptions.json во всех родительских папках"""
    # Начинаем с текущей папки и идем вверх до целевой директории
    current_dir = folder_path
    
    # Проверяем все уровни вверх
    while current_dir and os.path.exists(current_dir):
        json_path = os.path.join(current_dir, "descriptions.json")
        
        # Проверяем, существует ли уже JSON файл
        if not os.path.exists(json_path):
            # Создаем базовый JSON с title и subtitle
            descriptions = {}
            
            # Определяем title для текущей папки
            current_folder_name = os.path.basename(current_dir)
            
            if current_dir == folder_path:
                # Это целевая папка (подпапка или основная)
                if subfolder_name and os.path.basename(current_dir) == subfolder_name[:25].strip(" ."):
                    # Это подпапка
                    descriptions["__title__"] = folder_name if folder_name != "Неизвестная папка" else ""
                    descriptions["__subtitle__"] = subfolder_name if subfolder_name else ""
                else:
                    # Это основная папка
                    descriptions["__title__"] = folder_name if folder_name != "Неизвестная папка" else ""
                    descriptions["__subtitle__"] = ""
            else:
                # Это родительская папка
                descriptions["__title__"] = current_folder_name if current_folder_name != "Неизвестная папка" else ""
                descriptions["__subtitle__"] = ""
            
            try:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(descriptions, f, ensure_ascii=False, indent=4)
                logger.info(f"Создан descriptions.json в папке: {current_dir}")
                logger.info(f"Title: {descriptions['__title__']}, Subtitle: {descriptions['__subtitle__']}")
            except Exception as e:
                logger.error(f"Ошибка при создании JSON в {current_dir}: {e}")
        
        # Поднимаемся на уровень выше
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:  # Достигли корня
            break
        current_dir = parent_dir

def create_thumbnails_for_folder(folder_path, thumbnail_generator):
    """
    Создает миниатюры для всех изображений в указанной папке
    
    Args:
        folder_path (str): Путь к папке с изображениями
        thumbnail_generator: Экземпляр ThumbnailGenerator
    
    Returns:
        int: Количество созданных миниатюр
    """
    if not THUMBNAIL_GENERATOR_AVAILABLE:
        logger.warning("Генератор миниатюр недоступен, пропускаем создание миниатюр")
        return 0
    
    # Создаем папку для миниатюр
    thumbnails_folder = os.path.join(folder_path, "thumbnails")
    os.makedirs(thumbnails_folder, exist_ok=True)
    
    created_count = 0
    
    try:
        # Получаем список всех файлов в папке
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            
            # Проверяем, является ли файл изображением
            if os.path.isfile(file_path):
                ext = os.path.splitext(filename)[1].lower()
                image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp'}
                
                if ext in image_extensions:
                    # Создаем миниатюру
                    result = thumbnail_generator.createThumbnail(file_path, thumbnails_folder)
                    if result:
                        created_count += 1
                        logger.debug(f"Создана миниатюра для: {filename}")
        
        if created_count > 0:
            logger.info(f"Создано {created_count} миниатюр в {thumbnails_folder}")
        else:
            logger.info(f"В папке нет изображений для создания миниатюр: {folder_path}")
            
    except Exception as e:
        logger.error(f"Ошибка при создании миниатюр для папки {folder_path}: {e}")
    
    return created_count

def get_all_files_recursive(folder_path):
    """Получает все файлы из папки и всех ее подпапок"""
    all_files = {}
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # Полный путь к файлу
            full_path = os.path.join(root, file)
            # Относительный путь от исходной папки
            rel_path = os.path.relpath(full_path, folder_path)
            all_files[full_path] = {
                'rel_path': rel_path,
                'filename': file
            }
    
    return all_files

def DoFolder(source_folder, table, dest_folder):
    # Преобразуем относительные пути в абсолютные
    source_folder_fixed = resolve_relative_path(source_folder)
    table_fixed = resolve_relative_path(table)
    dest_folder_fixed = resolve_relative_path(dest_folder)
    
    # Логируем пути для отладки
    logger.info(f"Исходная папка: {source_folder_fixed}")
    logger.info(f"Excel файл: {table_fixed}")
    logger.info(f"Целевая папка: {dest_folder_fixed}")
    
    # Проверяем существование исходной папки
    if not os.path.exists(source_folder_fixed):
        logger.error(f"Исходная папка не найдена: {source_folder_fixed}")
        return
    
    # Проверяем существование Excel файла
    if not os.path.exists(table_fixed):
        logger.error(f"Excel файл не найден: {table_fixed}")
        return
    
    # Создаем конечную папку, если ее нет
    os.makedirs(dest_folder_fixed, exist_ok=True)
    
    # Создаем генератор миниатюр
    if THUMBNAIL_GENERATOR_AVAILABLE:
        thumbnail_generator = ThumbnailGenerator(
            max_size=THUMBNAIL_SIZE,
            quality=THUMBNAIL_QUALITY,
            format='JPEG'
        )
    else:
        thumbnail_generator = None
    
    # Создаем список для сбора информации о не найденных файлах
    not_found_reports = []
    
    # Загружаем все файлы из source_folder и всех подпапок
    logger.info(f"Поиск файлов в {source_folder_fixed} и подпапках...")
    all_files = get_all_files_recursive(source_folder_fixed)
    
    # Создаем карту нормализованных имен
    normalized_map = {}
    file_count = 0
    
    for full_path, file_info in all_files.items():
        filename = file_info['filename']
        normalized = normalize_filename(filename)
        
        if normalized not in normalized_map:
            normalized_map[normalized] = []
        
        normalized_map[normalized].append({
            'full_path': full_path,
            'filename': filename,
            'rel_path': file_info['rel_path']
        })
        file_count += 1
    
    # Логируем количество уникальных нормализованных имен
    logger.info(f"Найдено {file_count} файлов, {len(normalized_map)} уникальных нормализованных имен")
    
    # Загружаем Excel файл
    try:
        wb = load_workbook(table_fixed, data_only=True)
        logger.info(f"Загружен Excel файл с листами: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Ошибка при загрузке Excel файла: {e}")
        return
    
    # Обрабатываем каждый лист
    for sheet_name in wb.sheetnames:
        logger.info(f"Обработка листа: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Получаем название папки из первой строки, второго столбца (столбец B)        
        folder_name_cell = ws.cell(row=2, column=1).value
        if folder_name_cell:
            folder_name = str(folder_name_cell).strip()
            folder_name = sanitize_folder_name(folder_name)
        else:
            logger.error(f"Не найдена ячейка с заголовком на листе: {sheet_name}")
            folder_name = "Неизвестная папка"

        subfolder_name_cell = ws.cell(row=2, column=2).value
        if subfolder_name_cell:
            subfolder_name = str(subfolder_name_cell).strip()
            subfolder_name = sanitize_folder_name(subfolder_name)
        else:
            subfolder_name = ""

        logger.info(f"Создаю папку: {folder_name}")
        if subfolder_name:
            logger.info(f"Подпапка: {subfolder_name}")
        
        # Создаем папку для листа
        sheet_folder = os.path.join(dest_folder_fixed, folder_name)
        if subfolder_name != "":
            sheet_folder = os.path.join(sheet_folder, subfolder_name[:25].strip(" ."))
        os.makedirs(sheet_folder, exist_ok=True)

        # ОБЕСПЕЧИВАЕМ НАЛИЧИЕ descriptions.json ВО ВСЕХ РОДИТЕЛЬСКИХ ПАПКАХ
        ensure_descriptions_json_in_all_parents(sheet_folder, folder_name, subfolder_name)

        # Создаем словарь для описаний
        descriptions = {}
        title = folder_name
        if folder_name == "Неизвестная папка":
            title = ""
        
        descriptions["__title__"] = title 
        descriptions["__subtitle__"] = subfolder_name if subfolder_name else ""

        # Собираем статистику
        row_count = 0
        copied_count = 0
        not_found_count = 0
        multiple_found_count = 0
        
        # Список не найденных файлов для этого листа
        sheet_not_found = []
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Пропускаем пустые строки
            if not row or row[0] is None:
                continue
            
            image_name = str(row[0]).strip()
            description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            
            row_count += 1
            
            # Нормализуем имя для поиска
            normalized_search = normalize_filename(image_name)
            
            # Ищем совпадение файла
            if normalized_search in normalized_map:
                matched_files = normalized_map[normalized_search]
                
                if len(matched_files) > 1:
                    logger.warning(f"Найдено {len(matched_files)} файлов для '{image_name}': {[f['filename'] for f in matched_files]}")
                    multiple_found_count += 1
                
                # Берем первый найденный файл
                file_info = matched_files[0]
                matched_file_path = file_info['full_path']
                matched_filename = file_info['filename']
                
                # Копируем файл
                dest_path = os.path.join(sheet_folder, matched_filename)
                
                try:
                    # Проверяем, не существует ли уже файл
                    if os.path.exists(dest_path):
                        # Добавляем суффикс если файл уже существует
                        base, ext = os.path.splitext(matched_filename)
                        counter = 1
                        while os.path.exists(dest_path):
                            new_filename = f"{base}_{counter}{ext}"
                            dest_path = os.path.join(sheet_folder, new_filename)
                            counter += 1
                        
                        final_filename = os.path.basename(dest_path)
                        logger.info(f"Файл '{matched_filename}' уже существует, сохранен как '{final_filename}'")
                    else:
                        final_filename = matched_filename
                    
                    shutil.copy2(matched_file_path, dest_path)
                    
                    # Добавляем в словарь описаний
                    descriptions[final_filename] = description.split("\n")[0]
                    copied_count += 1
                    
                    logger.debug(f"Скопирован: {file_info['rel_path']} -> {folder_name}/{final_filename}")
                except Exception as e:
                    logger.error(f"Ошибка при копировании из {matched_file_path} в {dest_path}: {e}")
            else:
                logger.warning(f"Не найден файл для: '{image_name}' на листе '{sheet_name}'")
                
                # Сохраняем информацию о не найденном файле
                sheet_not_found.append({
                    'title': folder_name,
                    'subtitle': subfolder_name,
                    'name': image_name,
                    'sheet': sheet_name
                })
                
                # Для отладки: выводим похожие имена
                similar = [k for k in normalized_map.keys() if normalized_search in k or k in normalized_search]
                if similar:
                    logger.debug(f"Похожие нормализованные имена: {similar[:5]}")
                
                not_found_count += 1
        
        # Добавляем не найденные файлы из этого листа в общий отчет
        if sheet_not_found:
            not_found_reports.extend(sheet_not_found)
        
        # ВСЕГДА сохраняем descriptions.json, даже если нет изображений
        json_path = os.path.join(sheet_folder, "descriptions.json")
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(descriptions, f, ensure_ascii=False, indent=4)
            
            if copied_count > 0:
                logger.info(f"Создан {json_path} с {len(descriptions) - 2} записями файлов")
            else:
                logger.info(f"Создан {json_path} только с title и subtitle (нет изображений)")
            
            logger.info(f"Title: {descriptions['__title__']}, Subtitle: {descriptions['__subtitle__']}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении JSON: {e}")
        
        # СОЗДАЕМ МИНИАТЮРЫ ДЛЯ ЭТОЙ ПАПКИ
        if thumbnail_generator and os.path.exists(sheet_folder):
            thumbnails_created = create_thumbnails_for_folder(sheet_folder, thumbnail_generator)
            logger.info(f"Создано {thumbnails_created} миниатюр для папки: {sheet_folder}")
        
        logger.info(f"Лист '{sheet_name}': обработано {row_count} строк, найдено {copied_count}, не найдено {not_found_count}, множественные совпадения: {multiple_found_count}")
        
        # Сохраняем отчет о не найденных файлах для этого листа в его папке
        if not_found_count > 0:
            report_path = os.path.join(sheet_folder, "not_found_report.txt")
            try:
                with open(report_path, 'w', encoding='utf-8') as f:
                    f.write(f"Отчет о не найденных файлах\n")
                    f.write(f"============================\n\n")
                    f.write(f"Имя папки: {folder_name}\n")
                    f.write(f"Имя подпапки: {subfolder_name if subfolder_name else '(нет)'}\n")
                    f.write(f"Исходный лист: {sheet_name}\n")
                    f.write(f"Не найдено файлов: {not_found_count} из {row_count}\n\n")
                    f.write(f"Список не найденных файлов:\n")
                    f.write(f"--------------------------\n")
                    for item in sheet_not_found:
                        f.write(f"• {item['name']}\n")
            except Exception as e:
                logger.error(f"Ошибка при сохранении отчета: {e}")
    
    # Сохраняем общий отчет о не найденных файлах в корневой целевой папке
    if not_found_reports:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"not_found_report_{timestamp}.txt"
        report_path = os.path.join(dest_folder_fixed, report_filename)
        
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"ОТЧЕТ О НЕ НАЙДЕННЫХ ФАЙЛАХ\n")
                f.write(f"===========================\n\n")
                f.write(f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
                f.write(f"Исходная папка: {source_folder_fixed}\n")
                f.write(f"Excel файл: {table_fixed}\n")
                f.write(f"Целевая папка: {dest_folder_fixed}\n")
                f.write(f"Всего не найденных файлов: {len(not_found_reports)}\n\n")
                f.write(f"СПИСОК НЕ НАЙДЕННЫХ ФАЙЛОВ:\n")
                f.write(f"============================\n\n")
                
                for item in not_found_reports:
                    f.write(f"Заголовок: {item['title']}\n")
                    f.write(f"Подзаголовок: {item['subtitle']}\n")
                    f.write(f"Название файла: {item['name']}\n")
                    f.write(f"Лист Excel: {item['sheet']}\n")
                    f.write(f"{'-'*50}\n")
            
            logger.info(f"Создан общий отчет о не найденных файлов: {report_path}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении общего отчета: {e}")
    else:
        logger.info("Все файлы успешно найдены. Отчет не требуется.")
    
    logger.info("Обработка завершена!")
    
    # Дополнительная диагностика
    logger.info("=== ДИАГНОСТИКА ===")
    logger.info(f"Всего файлов в исходной папке: {file_count}")
    logger.info(f"Уникальных нормализованных имен: {len(normalized_map)}")

def main():
    # Используем только относительные пути
    # Все пути будут автоматически преобразованы в абсолютные
    # относительно директории скрипта
    
    DoFolder(
        r"Unsorted\Peterhof\до войны",
        r"Unsorted\Peterhof\до войны\!до_войны_подписи_текст4.xlsx",
        r"Sorted\Peterhof\до войны"
    )
    logger.debug(f"##################################################################################################################")

    DoFolder(
        r"Unsorted\Peterhof\разрушения",
        r"Unsorted\Peterhof\разрушения\!разрушения_подписи_текст4.xlsx",
        r"Sorted\Peterhof\разрушения"
    )
    logger.debug(f"##################################################################################################################")
    
    DoFolder(
        r"Unsorted\Peterhof\восстановление",
        r"Unsorted\Peterhof\восстановление\!восстановление_подписи_текст4.xlsx",
        r"Sorted\Peterhof\восстановление"
    )

#Pushkin
    DoFolder(
        r"Unsorted\Pushkin\до войны",
        r"Unsorted\Pushkin\до войны\до войны_подписи_пушкин.xlsx",
        r"Sorted\Pushkin\до войны"
    )
    logger.debug(f"##################################################################################################################")

    DoFolder(
        r"Unsorted\Pushkin\разрушения",
        r"Unsorted\Pushkin\разрушения\разрушения_подписи_пушкин.xlsx",
        r"Sorted\Pushkin\разрушения"
    )
    logger.debug(f"##################################################################################################################")
    DoFolder(
        r"Unsorted\Pushkin\восстановление",
        r"Unsorted\Pushkin\восстановление\восстановление_подписи_пушкин.xlsx",
        r"Sorted\Pushkin\восстановление"
    )
#Pavlovsk
    DoFolder(
        r"Unsorted\Pavlovsk\до войны",
        r"Unsorted\Pavlovsk\до войны\до войны_павловск_подписи.xlsx",
        r"Sorted\Pavlovsk\до войны"
    )
    logger.debug(f"##################################################################################################################")

    DoFolder(
        r"Unsorted\Pavlovsk\разрушения",
        r"Unsorted\Pavlovsk\разрушения\разрушение_павловск_подписи.xlsx",
        r"Sorted\Pavlovsk\разрушения"
    )
    logger.debug(f"##################################################################################################################")
    DoFolder(
        r"Unsorted\Pavlovsk\восстановление",
        r"Unsorted\Pavlovsk\восстановление\восстановление_павловск_подписи.xlsx",
        r"Sorted\Pavlovsk\восстановление"
    )

if __name__ == "__main__":
    main()