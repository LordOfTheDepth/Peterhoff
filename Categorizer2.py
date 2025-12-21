import os
import json
import re
import shutil
from openpyxl import load_workbook
from pathlib import Path
import logging

# ========== НАСТРОЙКИ ==========

# ===============================

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def normalize_filename(filename):
    """Удаляет расширения изображений и оставляет только буквы и цифры"""
    # Определяем расширения файлов изображений
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg'}
    
    # Находим расширение файла (последняя точка в имени)
    name, ext = os.path.splitext(filename)
    ext_lower = ext.lower()
    
    # Если это расширение изображения - удаляем его, иначе оставляем как есть
    if ext_lower in image_extensions:
        name_without_ext = name
    else:
        name_without_ext = filename  # Если не изображение, оставляем все имя
    
    # Создаем пустую строку для результата
    normalized_chars = []
    
    # Проходим по каждому символу
    for char in name_without_ext:
        if char.isalpha() or char.isdigit():
            normalized_chars.append(char)
    
    # Объединяем и приводим к нижнему регистру
    normalized = ''.join(normalized_chars)
    return normalized.lower()

def sanitize_folder_name(folder_name):
    """Очищает имя папки от недопустимых символов"""
    # Удаляем недопустимые символы для имен папок в Windows
    # <>:"/\|?* а также управляющие символы
    invalid_chars = '<>:"/\\|?*\n'
    for char in invalid_chars:
        folder_name = folder_name.replace(char, '')
    

    # Ограничиваем длину (опционально)

        
    # Удаляем начальные и конечные пробелы и точки
    folder_name = folder_name.strip(' .')
    
    
    # Если после очистки имя пустое, используем имя листа
    if not folder_name:
        return "Без_названия"
    
    return folder_name

def get_all_files_recursive(folder_path):
    """Получает все файлы из папки и всех ее подпапок"""
    all_files = {}
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # Полный путь к файлу
            full_path = os.path.join(root, file)
            # Относительный путь от исходной папки
            rel_path = os.path.relpath(full_path, folder_path)
            all_files[full_path] = {
                'rel_path': rel_path,
                'filename': file
            }
    
    return all_files

def DoFolder(source_folder, table, dest_folder):
    # Используем raw strings для путей в Windows
    source_folder_fixed = source_folder
    table_fixed = table
    dest_folder_fixed = dest_folder
    # Проверяем существование исходной папки
    if not os.path.exists(source_folder_fixed):
        logger.error(f"Исходная папка не найдена: {source_folder_fixed}")
        return
    
    # Проверяем существование Excel файла
    if not os.path.exists(table_fixed):
        logger.error(f"Excel файл не найден: {table_fixed}")
        return
    
    # Создаем конечную папку, если ее нет
    os.makedirs(dest_folder_fixed, exist_ok=True)
    
    # Загружаем все файлы из source_folder и всех подпапок
    logger.info(f"Поиск файлов в {source_folder_fixed} и подпапках...")
    all_files = get_all_files_recursive(source_folder_fixed)
    
    # Создаем карту нормализованных имен
    # Ключ: нормализованное имя, Значение: список путей к файлам с таким именем
    normalized_map = {}
    file_count = 0
    
    for full_path, file_info in all_files.items():
        filename = file_info['filename']
        normalized = normalize_filename(filename)
        
        if normalized not in normalized_map:
            normalized_map[normalized] = []
        
        normalized_map[normalized].append({
            'full_path': full_path,
            'filename': filename,
            'rel_path': file_info['rel_path']
        })
        file_count += 1
    
    # Логируем количество уникальных нормализованных имен
    logger.info(f"Найдено {file_count} файлов, {len(normalized_map)} уникальных нормализованных имен")
    
    # Выводим несколько примеров для отладки
    logger.debug(f"Примеры нормализованных имен (первые 5): {list(normalized_map.keys())[:5]}")
    
    # Загружаем Excel файл
    try:
        wb = load_workbook(table_fixed, data_only=True)
        logger.info(f"Загружен Excel файл с листами: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Ошибка при загрузке Excel файла: {e}")
        return
    
    # Обрабатываем каждый лист
    for sheet_name in wb.sheetnames:
        logger.info(f"Обработка листа: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Получаем название папки из первой строки, второго столбца (столбец B)        
        folder_name_cell = ws.cell(row=2, column=1).value
        if folder_name_cell:
            folder_name = str(folder_name_cell).strip()
            folder_name = sanitize_folder_name(folder_name)
        else:
            logger.error(f"Не найдена ячейка с заголовком на листе: {sheet_name}")
            folder_name = "Неизвестная папка"

        subfolder_name_cell = ws.cell(row=2, column=2).value
        if subfolder_name_cell:
            subfolder_name = str(subfolder_name_cell).strip()
            subfolder_name = sanitize_folder_name(subfolder_name)
        else:
            subfolder_name = ""

        logger.info(f"Создаю папку: {folder_name}")
        if subfolder_name:
            logger.info(f"Подпапка: {subfolder_name}")
        
        # Создаем папку для листа
        sheet_folder = os.path.join(dest_folder_fixed, folder_name)
        if subfolder_name != "":
            sheet_folder = os.path.join(sheet_folder, subfolder_name[:25].strip(" ."))
        os.makedirs(sheet_folder, exist_ok=True)

        # Создаем словарь для описаний
        # Сохраняем старую структуру, но добавляем title и subtitle как первые ключи
        descriptions = {}
        descriptions["__title__"] = folder_name
        descriptions["__subtitle__"] = subfolder_name if subfolder_name else ""

        # Собираем статистику
        row_count = 0
        copied_count = 0
        not_found_count = 0
        multiple_found_count = 0
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Пропускаем пустые строки
            if not row or row[0] is None:
                continue
            
            image_name = str(row[0]).strip()
            description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            
            row_count += 1
            
            # Нормализуем имя для поиска
            normalized_search = normalize_filename(image_name)
            
            # Ищем совпадение файла
            if normalized_search in normalized_map:
                matched_files = normalized_map[normalized_search]
                
                if len(matched_files) > 1:
                    logger.warning(f"Найдено {len(matched_files)} файлов для '{image_name}': {[f['filename'] for f in matched_files]}")
                    multiple_found_count += 1
                
                # Берем первый найденный файл (можно изменить логику если нужно все)
                file_info = matched_files[0]
                matched_file_path = file_info['full_path']
                matched_filename = file_info['filename']
                
                # Копируем файл
                dest_path = os.path.join(sheet_folder, matched_filename)
                
                try:
                    # Проверяем, не существует ли уже файл
                    if os.path.exists(dest_path):
                        # Добавляем суффикс если файл уже существует
                        base, ext = os.path.splitext(matched_filename)
                        counter = 1
                        while os.path.exists(dest_path):
                            new_filename = f"{base}_{counter}{ext}"
                            dest_path = os.path.join(sheet_folder, new_filename)
                            counter += 1
                        
                        final_filename = os.path.basename(dest_path)
                        logger.info(f"Файл '{matched_filename}' уже существует, сохранен как '{final_filename}'")
                    else:
                        final_filename = matched_filename
                    
                    shutil.copy2(matched_file_path, dest_path)
                    
                    # Добавляем в словарь описаний (старая структура)
                    descriptions[final_filename] = description
                    copied_count += 1
                    
                    logger.debug(f"Скопирован: {file_info['rel_path']} -> {folder_name}/{final_filename}")
                except Exception as e:
                    logger.error(f"Ошибка при копировании из {matched_file_path} в {dest_path}: {e}")
            else:
                logger.warning(f"Не найден файл для: '{image_name}' на листе '{sheet_name}'")
                
                # Для отладки: выводим похожие имена
                similar = [k for k in normalized_map.keys() if normalized_search in k or k in normalized_search]
                if similar:
                    logger.debug(f"Похожие нормализованные имена: {similar[:5]}")
                
                not_found_count += 1
        
        # Сохраняем descriptions.json
        json_path = os.path.join(sheet_folder, "descriptions.json")
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(descriptions, f, ensure_ascii=False, indent=4)
            logger.info(f"Создан {json_path} с {len(descriptions) - 2} записями файлов")  # -2 потому что title и subtitle
            logger.info(f"Title: {descriptions['__title__']}, Subtitle: {descriptions['__subtitle__']}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении JSON: {e}")
        
        logger.info(f"Лист '{sheet_name}': обработано {row_count} строк, найдено {copied_count}, не найдено {not_found_count}, множественные совпадения: {multiple_found_count}")
        
        # Сохраняем отчет о не найденных файлах
        if not_found_count > 0:
            report_path = os.path.join(sheet_folder, "not_found_report.txt")
            try:
                with open(report_path, 'w', encoding='utf-8') as f:
                    f.write(f"Не найдено файлов: {not_found_count} из {row_count}\n")
                    f.write(f"Имя папки: {folder_name}\n")
                    f.write(f"Имя подпапки: {subfolder_name if subfolder_name else '(нет)'}\n")
                    f.write(f"Исходный лист: {sheet_name}\n")
            except Exception as e:
                logger.error(f"Ошибка при сохранении отчета: {e}")
    
    logger.info("Обработка завершена!")
    
    # Дополнительная диагностика
    logger.info("=== ДИАГНОСТИКА ===")
    logger.info(f"Всего файлов в исходной папке: {file_count}")
    logger.info(f"Уникальных нормализованных имен: {len(normalized_map)}")

def main():

    mainFolder = r"C:\liferay-ce-portal-7.2.1-ga2\PeterhoffParts\Peterhoff\Sorted"
        
    DoFolder(
        r"C:\Peterhof\до войны",
        r"C:\Peterhof\до войны\!до_войны_подписи_текст3.xlsx",
        r"C:\liferay-ce-portal-7.2.1-ga2\PeterhoffParts\Peterhoff\Sorted\до войны"
    )
    logger.debug(f"##################################################################################################################")

    DoFolder(
        r"C:\Peterhof\разрушения",
        r"C:\Peterhof\разрушения\!разрушения_подписи_текст3.xlsx",
        r"C:\liferay-ce-portal-7.2.1-ga2\PeterhoffParts\Peterhoff\Sorted\разрушения"
    )
    logger.debug(f"##################################################################################################################")
    DoFolder(
        r"C:\Peterhof\восстановление",
        r"C:\Peterhof\восстановление\!восстановление_подписи_текст3.xlsx",
        r"C:\liferay-ce-portal-7.2.1-ga2\PeterhoffParts\Peterhoff\Sorted\восстановление"
    )


if __name__ == "__main__":
    main()